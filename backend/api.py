"""
backend/api.py — Rotas da API REST do Hub de Automações.
"""

import csv
import io
import json
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, Response, StreamingResponse

router = APIRouter()

_DB_CONTRATACOES = Path("database/contratacoes.json")
_DB_HISTORICO    = Path("database/historico.json")

_VAZIO = json.dumps(
    {"ultima_atualizacao": None, "total": 0, "chamados": []},
    ensure_ascii=False,
)

_VAZIO_HIST = json.dumps([], ensure_ascii=False)


@router.get("/contratacoes", summary="Chamados de contratação ativos")
async def get_contratacoes() -> Response:
    if not _DB_CONTRATACOES.exists():
        return Response(content=_VAZIO, media_type="application/json")
    try:
        return Response(
            content=_DB_CONTRATACOES.read_text(encoding="utf-8"),
            media_type="application/json",
        )
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"erro": str(exc), "total": 0, "chamados": []},
        )


@router.get("/historico", summary="KPIs históricos por dia")
async def get_historico() -> Response:
    if not _DB_HISTORICO.exists():
        return Response(content=_VAZIO_HIST, media_type="application/json")
    try:
        return Response(
            content=_DB_HISTORICO.read_text(encoding="utf-8"),
            media_type="application/json",
        )
    except Exception as exc:
        return JSONResponse(status_code=500, content={"erro": str(exc)})


@router.get("/contratacoes/export", summary="Exporta chamados em CSV ou XLSX")
async def export_contratacoes(
    formato: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    termo: str | None = Query(None),
) -> StreamingResponse:
    if not _DB_CONTRATACOES.exists():
        return JSONResponse(status_code=404, content={"erro": "Dados ainda não disponíveis."})

    data = json.loads(_DB_CONTRATACOES.read_text(encoding="utf-8"))
    chamados: list[dict] = data.get("chamados", [])

    if termo:
        chamados = [c for c in chamados if c.get("Termo_Status") == termo]

    if formato == "csv":
        return _export_csv(chamados, data.get("ultima_atualizacao"))
    return _export_xlsx(chamados, data.get("ultima_atualizacao"))


# ── helpers de export ────────────────────────────────────────────────────────

_COLUNAS = ["ID_do_Chamado", "Titulo", "Status", "Tempo_Solucao",
            "Data_Abertura", "Requerente", "Termo_Status"]

_HEADERS_PT = {
    "ID_do_Chamado": "ID",
    "Titulo":        "Título",
    "Status":        "Status",
    "Tempo_Solucao": "Tempo p/ Solução",
    "Data_Abertura": "Data de Abertura",
    "Requerente":    "Requerente",
    "Termo_Status":  "Status do Termo",
}


def _ts_filename(ultima_atualizacao: str | None) -> str:
    try:
        d = datetime.fromisoformat(ultima_atualizacao or "")
        return d.strftime("%Y%m%d_%H%M")
    except Exception:
        return datetime.now().strftime("%Y%m%d_%H%M")


def _export_csv(chamados: list[dict], ultima_atualizacao: str | None) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=_COLUNAS,
        extrasaction="ignore",
        lineterminator="\r\n",
    )
    writer.writerow({c: _HEADERS_PT[c] for c in _COLUNAS})
    writer.writerows(chamados)
    ts = _ts_filename(ultima_atualizacao)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="contratacoes_{ts}.csv"'},
    )


def _export_xlsx(chamados: list[dict], ultima_atualizacao: str | None) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                     Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse(
            status_code=500,
            content={"erro": "openpyxl não instalado. Execute: pip install openpyxl"},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratações"

    # ── Cores e estilos ───────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")   # azul-marinho
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)

    THIN  = Side(style="thin",   color="D1D5DB")
    THICK = Side(style="medium", color="1E3A5F")
    BORDER_HDR  = Border(bottom=Side(style="medium", color="4F46E5"))
    BORDER_CELL = Border(
        left=THIN, right=THIN, top=THIN, bottom=THIN,
    )

    TERMO_FILLS = {
        "Termo OK":          PatternFill("solid", fgColor="D1FAE5"),
        "Pendente":          PatternFill("solid", fgColor="FEF3C7"),
        "Erro ao verificar": PatternFill("solid", fgColor="FFE4E6"),
        "Sem tarefa":        PatternFill("solid", fgColor="FFE4E6"),
    }
    TERMO_FONTS = {
        "Termo OK":          Font(color="065F46", size=10, bold=True),
        "Pendente":          Font(color="92400E", size=10, bold=True),
        "Erro ao verificar": Font(color="9F1239", size=10, bold=True),
        "Sem tarefa":        Font(color="9F1239", size=10, bold=True),
    }
    ROW_FILL_A = PatternFill("solid", fgColor="F8FAFC")
    ROW_FILL_B = PatternFill("solid", fgColor="FFFFFF")
    CELL_FONT  = Font(size=10, color="1E293B")
    CELL_ALIGN = Alignment(vertical="center")

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    for col_idx, col_key in enumerate(_COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADERS_PT[col_key])
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BORDER_HDR

    # ── Dados ─────────────────────────────────────────────────────────────────
    for row_idx, chamado in enumerate(chamados, start=2):
        ws.row_dimensions[row_idx].height = 22
        fill_base = ROW_FILL_A if row_idx % 2 == 0 else ROW_FILL_B

        for col_idx, col_key in enumerate(_COLUNAS, start=1):
            val  = chamado.get(col_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER_CELL
            cell.alignment = CELL_ALIGN

            if col_key == "Termo_Status":
                cell.fill = TERMO_FILLS.get(val, fill_base)
                cell.font = TERMO_FONTS.get(val, CELL_FONT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_key == "ID_do_Chamado":
                cell.fill = fill_base
                cell.font = Font(size=10, color="6366F1", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_key == "Status":
                cell.fill = fill_base
                cell.font = Font(size=10, color="374151", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = fill_base
                cell.font = CELL_FONT

    # ── Largura automática ────────────────────────────────────────────────────
    MIN_W = {"ID_do_Chamado": 8, "Status": 16, "Termo_Status": 18,
             "Data_Abertura": 18, "Tempo_Solucao": 20}
    MAX_W = {"Titulo": 52, "Requerente": 30}
    for col_idx, col_key in enumerate(_COLUNAS, start=1):
        col_letter = get_column_letter(col_idx)
        values = [_HEADERS_PT[col_key]] + [str(c.get(col_key, "") or "") for c in chamados]
        best_w = max(len(v) for v in values) + 2
        best_w = max(best_w, MIN_W.get(col_key, 12))
        best_w = min(best_w, MAX_W.get(col_key, 60))
        ws.column_dimensions[col_letter].width = best_w

    # ── Freeze pane + auto-filter ──────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS))}1"

    # ── Rodapé de metadados ───────────────────────────────────────────────────
    footer_row = len(chamados) + 3
    ws.row_dimensions[footer_row].height = 16
    ts_str = ""
    try:
        d = datetime.fromisoformat(ultima_atualizacao or "")
        ts_str = d.strftime("%d/%m/%Y %H:%M")
    except Exception:
        ts_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    meta_cell = ws.cell(
        row=footer_row, column=1,
        value=f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Snapshot GLPI: {ts_str} | Total: {len(chamados)} chamado(s)",
    )
    meta_cell.font      = Font(size=9, color="94A3B8", italic=True)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=footer_row, start_column=1,
        end_row=footer_row,   end_column=len(_COLUNAS),
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = _ts_filename(ultima_atualizacao)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contratacoes_{ts}.xlsx"'},
    )
